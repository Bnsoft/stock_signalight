"""Data Export - Export data to CSV, PDF, and Excel formats"""

from datetime import datetime, timedelta
from typing import List, Dict, Optional, BinaryIO
import csv
import json
from io import StringIO, BytesIO
from . import store


# ============= CSV Export =============

def export_portfolio_to_csv(user_id: str) -> str:
    """Export portfolio to CSV"""
    with store._connect() as conn:
        positions = conn.execute(
            """SELECT symbol, quantity, entry_price, current_price, position_value,
                      position_pnl, position_return
               FROM positions
               WHERE user_id = ?""",
            (user_id,),
        ).fetchall()

    csv_buffer = StringIO()
    writer = csv.writer(csv_buffer)

    # Header
    writer.writerow([
        "종목",
        "수량",
        "진입가",
        "현재가",
        "포지션 가치",
        "손익",
        "수익률(%)",
        "내보내기 일자",
    ])

    # Data
    total_value = 0
    total_pnl = 0

    for pos in positions:
        writer.writerow([
            pos[0],  # symbol
            pos[1],  # quantity
            f"${pos[2]:.2f}",  # entry_price
            f"${pos[3]:.2f}",  # current_price
            f"${pos[4]:.2f}",  # position_value
            f"${pos[5]:.2f}",  # position_pnl
            f"{pos[6]:.2f}%",  # position_return
            datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S"),
        ])
        total_value += pos[4]
        total_pnl += pos[5]

    # Summary
    writer.writerow([])
    writer.writerow(["합계", "", "", "", f"${total_value:.2f}", f"${total_pnl:.2f}"])

    return csv_buffer.getvalue()


def export_backtest_results_to_csv(
    symbol: str,
    strategy: str,
    backtest_data: Dict,
) -> str:
    """Export backtest results to CSV"""
    csv_buffer = StringIO()
    writer = csv.writer(csv_buffer)

    # Header
    writer.writerow(["백테스트 결과 리포트"])
    writer.writerow([f"종목: {symbol}", f"전략: {strategy}"])
    writer.writerow([f"생성일: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')}"])
    writer.writerow([])

    # Key metrics
    writer.writerow(["지표", "값"])
    writer.writerow(["초기 자본", f"${backtest_data.get('initial_capital', 0):.2f}"])
    writer.writerow(["최종 자본", f"${backtest_data.get('final_value', 0):.2f}"])
    writer.writerow(["총 수익률", f"{backtest_data.get('total_return_percent', 0):.2f}%"])
    writer.writerow(["총 거래 수", backtest_data.get("total_trades", 0)])
    writer.writerow(["수익 거래", backtest_data.get("winning_trades", 0)])
    writer.writerow(["손실 거래", backtest_data.get("losing_trades", 0)])
    writer.writerow(["승률", f"{backtest_data.get('win_rate', 0):.2f}%"])
    writer.writerow(["Sharpe Ratio", f"{backtest_data.get('sharpe_ratio', 0):.2f}"])
    writer.writerow(["최대 손실률", f"{backtest_data.get('max_drawdown_percent', 0):.2f}%"])

    return csv_buffer.getvalue()


def export_alerts_to_csv(user_id: str) -> str:
    """Export alerts to CSV"""
    with store._connect() as conn:
        alerts = conn.execute(
            """SELECT id, symbol, alert_type, trigger_price, notify_methods, status, created_at
               FROM price_alerts
               WHERE user_id = ?
               UNION ALL
               SELECT id, symbol, indicator as alert_type, threshold as trigger_price,
                      notify_methods, status, created_at
               FROM indicator_alerts
               WHERE user_id = ?""",
            (user_id, user_id),
        ).fetchall()

    csv_buffer = StringIO()
    writer = csv.writer(csv_buffer)

    # Header
    writer.writerow([
        "알람 ID",
        "종목",
        "알람 타입",
        "트리거 가격/조건",
        "알림 방식",
        "상태",
        "생성 일자",
    ])

    # Data
    for alert in alerts:
        writer.writerow([
            alert[0],
            alert[1],
            alert[2],
            alert[3],
            alert[4],
            alert[5],
            alert[6],
        ])

    return csv_buffer.getvalue()


def export_transactions_to_csv(user_id: str, start_date: Optional[str] = None) -> str:
    """Export transaction history to CSV"""
    with store._connect() as conn:
        query = """SELECT symbol, order_type, quantity, price, created_at, status
                   FROM broker_orders
                   WHERE user_id = ?"""
        params = [user_id]

        if start_date:
            query += " AND created_at >= ?"
            params.append(start_date)

        transactions = conn.execute(query, params).fetchall()

    csv_buffer = StringIO()
    writer = csv.writer(csv_buffer)

    # Header
    writer.writerow([
        "종목",
        "주문 타입",
        "수량",
        "가격",
        "거래 일시",
        "상태",
    ])

    # Data
    for trans in transactions:
        writer.writerow([
            trans[0],
            trans[1],
            trans[2],
            f"${trans[3]:.2f}" if trans[3] else "N/A",
            trans[4],
            trans[5],
        ])

    return csv_buffer.getvalue()


# ============= Excel Export =============

def export_portfolio_to_excel(user_id: str) -> bytes:
    """Export portfolio to Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        # Fall back to CSV if openpyxl is not available
        return export_portfolio_to_csv(user_id).encode('utf-8')

    with store._connect() as conn:
        positions = conn.execute(
            """SELECT symbol, quantity, entry_price, current_price, position_value,
                      position_pnl, position_return
               FROM positions
               WHERE user_id = ?""",
            (user_id,),
        ).fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "포트폴리오"

    # Header
    headers = ["종목", "수량", "진입가", "현재가", "포지션 가치", "손익", "수익률(%)"]
    ws.append(headers)

    # Header style
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Data
    total_value = 0
    total_pnl = 0

    for pos in positions:
        ws.append([
            pos[0],  # symbol
            pos[1],  # quantity
            pos[2],  # entry_price
            pos[3],  # current_price
            pos[4],  # position_value
            pos[5],  # position_pnl
            pos[6],  # position_return
        ])
        total_value += pos[4]
        total_pnl += pos[5]

    # Summary row
    ws.append([])
    ws.append(["합계", "", "", "", total_value, total_pnl, ""])

    # Column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 15
    ws.column_dimensions["G"].width = 12

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


# ============= PDF Export =============

def export_portfolio_to_pdf(user_id: str) -> bytes:
    """Export portfolio to PDF"""
    try:
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.lib import colors
    except ImportError:
        # Fall back to CSV if reportlab is not available
        return export_portfolio_to_csv(user_id).encode('utf-8')

    with store._connect() as conn:
        positions = conn.execute(
            """SELECT symbol, quantity, entry_price, current_price, position_value,
                      position_pnl, position_return
               FROM positions
               WHERE user_id = ?""",
            (user_id,),
        ).fetchall()

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    story = []

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "CustomTitle",
        parent=styles["Heading1"],
        fontSize=24,
        textColor=colors.HexColor("#1f3a93"),
        spaceAfter=30,
    )

    # Title
    story.append(Paragraph("포트폴리오 리포트", title_style))
    story.append(Paragraph(f"생성 일자: {datetime.utcnow().strftime('%Y년 %m월 %d일')}", styles["Normal"]))
    story.append(Spacer(1, 0.2 * inch))

    # Table data
    data = [["종목", "수량", "진입가", "현재가", "포지션 가치", "손익", "수익률(%)"]]

    total_value = 0
    total_pnl = 0

    for pos in positions:
        data.append([
            pos[0],
            str(pos[1]),
            f"${pos[2]:.2f}",
            f"${pos[3]:.2f}",
            f"${pos[4]:.2f}",
            f"${pos[5]:.2f}",
            f"{pos[6]:.2f}%",
        ])
        total_value += pos[4]
        total_pnl += pos[5]

    data.append(["합계", "", "", "", f"${total_value:.2f}", f"${total_pnl:.2f}", ""])

    # Create table
    table = Table(data)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 12),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#D3D3D3")),
        ("GRID", (0, 0), (-1, -1), 1, colors.black),
    ]))

    story.append(table)

    doc.build(story)
    return buffer.getvalue()


# ============= Report Generation =============

def generate_monthly_report(user_id: str, year: int, month: int) -> Dict:
    """Generate monthly report"""
    start_date = f"{year}-{month:02d}-01"
    end_date = f"{year}-{month:02d}-28"

    with store._connect() as conn:
        # Fetch trades
        trades = conn.execute(
            """SELECT symbol, quantity, price, created_at FROM broker_orders
               WHERE user_id = ? AND created_at BETWEEN ? AND ? AND status = 'FILLED'
               ORDER BY created_at""",
            (user_id, start_date, end_date),
        ).fetchall()

        # Portfolio state
        positions = conn.execute(
            """SELECT symbol, quantity, position_value, position_pnl, position_return
               FROM positions WHERE user_id = ?""",
            (user_id,),
        ).fetchall()

    total_pnl = sum(p[3] for p in positions)
    total_value = sum(p[2] for p in positions)
    monthly_return = (total_pnl / total_value * 100) if total_value > 0 else 0

    return {
        "year": year,
        "month": month,
        "period": f"{year}-{month:02d}",
        "trades_count": len(trades),
        "portfolio_value": round(total_value, 2),
        "monthly_pnl": round(total_pnl, 2),
        "monthly_return_percent": round(monthly_return, 2),
        "positions_count": len(positions),
        "top_position": max(positions, key=lambda x: x[2])[0] if positions else "N/A",
    }


def generate_annual_report(user_id: str, year: int) -> Dict:
    """Generate annual report"""
    start_date = f"{year}-01-01"
    end_date = f"{year}-12-31"

    monthly_reports = []
    total_annual_pnl = 0

    for month in range(1, 13):
        report = generate_monthly_report(user_id, year, month)
        monthly_reports.append(report)
        total_annual_pnl += report["monthly_pnl"]

    with store._connect() as conn:
        trades = conn.execute(
            """SELECT COUNT(*) FROM broker_orders
               WHERE user_id = ? AND created_at BETWEEN ? AND ? AND status = 'FILLED'""",
            (user_id, start_date, end_date),
        ).fetchone()

        positions = conn.execute(
            """SELECT SUM(position_pnl), SUM(position_value)
               FROM positions WHERE user_id = ?""",
            (user_id,),
        ).fetchone()

    total_value = positions[1] or 0
    annual_return = (total_annual_pnl / total_value * 100) if total_value > 0 else 0

    return {
        "year": year,
        "total_trades": trades[0],
        "total_annual_pnl": round(total_annual_pnl, 2),
        "annual_return_percent": round(annual_return, 2),
        "best_month": max(monthly_reports, key=lambda x: x["monthly_return_percent"])["period"],
        "worst_month": min(monthly_reports, key=lambda x: x["monthly_return_percent"])["period"],
        "monthly_breakdown": monthly_reports,
    }


def export_report_to_json(report: Dict) -> str:
    """Export report to JSON"""
    return json.dumps(report, indent=2, ensure_ascii=False, default=str)


def create_export_summary(user_id: str) -> Dict:
    """Export summary"""
    return {
        "user_id": user_id,
        "export_timestamp": datetime.utcnow().isoformat(),
        "formats_available": ["CSV", "Excel", "PDF", "JSON"],
        "data_types": [
            "Portfolio",
            "Backtest Results",
            "Alerts",
            "Transactions",
            "Monthly Report",
            "Annual Report",
        ],
        "notes": "모든 데이터는 안전하게 내보내집니다",
    }
